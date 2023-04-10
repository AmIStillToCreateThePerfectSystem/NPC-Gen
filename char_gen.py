# -*- coding: utf-8 -*-
"""
Created on Tue Oct 11 00:57:52 2022

"""
# %% Modules
import pandas as pd
import openpyxl
import sys
import os
import numpy as np
import random
import time
print("Importing modules...")

# import xlsxwriter
# import xlrd
# from openpyxl.utils import get_column_letter
# from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
# from inspect import currentframe, getframeinfo
# import copy

# %% Functions
print("Loading functions...")


def lower_bound(search_list, result_list, x):
    if not search_list or search_list[0] > x:
        return
    for i, y in enumerate(search_list):
        if y > x:
            return result_list[i - 1]


def check_lowest(search_list, result_list, x):
    result = 0

    for i, s in enumerate(search_list):
        if x < s:
            result = i - 1 if i > 0 else 0
            break
        elif x >= search_list[-1]:  # to catch the highest roll
            result = len(search_list) - 1

    # print(f"Roll : {x}, Result : {result_list[result]}")
    return result_list[result]


def population_input(race):
    population = input(f"What is the {race} population of this area? (In families) : ")

    try:
        total_pop = int(population)*5
        # print(f"Total population is {total_pop}")
        return total_pop
    except:
        print("Not a valid number.")
        population_input(race)


def xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return openpyxl.get_column_letter(column) + str(row)


def excel_dict_fetch(file, named_range):
    dest = file.defined_names[named_range].destinations
    for title, coord in dest:
        data = file[title][coord]
    raw = data.value
    return dict(eval(raw))


def stat_assign(current_char):
    char_stats = current_char["Stats"]
    race = current_char.get("Race")
    # print("Current stat id : ",id(current_stats))
    for stat in stats:
        roll_type = new_race_info_dict["Stat Prob"][race][stat]
        stat_index = stat_index_all[roll_type]
        char_stats[stat] = int(stat_rolls[roll_type][stat_index])
        stat_index += 1
        stat_index_all[roll_type] = stat_index
    current_char["Stats"] = char_stats
    # print(char_stats)
    # print(stat_index_all)
    return current_char


def stat_weighting(current_stats, primes):
    sum_of_values = sum([max(0, i[0]-12) for i in zip(current_stats, primes) if i[1]])
    formula = 1.6 ** sum_of_values
    weight = int(round(formula, 0))
    return weight


def class_assign(current_char):
    # current_char = characters[1]
    total_weight = 0
    char_class_weights = {}
    race = current_char.get("Race")
    for current_class in new_race_info_dict["Class"][race]:
        # current_class = "Fighter"
        current_stats = [current_char["Stats"][i] for i in stats]
        stat_minimums = [new_class_stat_dict[current_class][i+"Min"] for i in stats]
        stat_check = [i[0] >= i[1] for i in zip(current_stats, stat_minimums)]
        if all(stat_check):
            primes = [new_class_stat_dict[current_class][i+"Prime"] for i in stats]
            # stat_weight = sum([stat_weights[i[0]] for i in zip(current_stats,primes) if i[1]])
            stat_weight = stat_weighting(current_stats, primes)
            base_weight = new_race_info_dict["Class"][race][current_class]
            current_weight = base_weight * stat_weight
        else:
            current_weight = 0
        # print(f"Line Number : {getframeinfo(currentframe()).lineno}\nStats : {attributes['Weights']}")
        char_class_weights[current_class] = current_weight
        total_weight += current_weight
    if total_weight > 0:
        new_weights = [char_class_weights[i]/total_weight for i in char_class_weights]
        # class_weight_prob_dict = {list(char_class_weights.keys())[i] : new_weights[i] for i in range(len(new_weights))}
        class_selection = [key for key in char_class_weights]
        new_class = str(np.random.choice(class_selection, 1, p=new_weights)[0])
        new_template = str(np.random.choice(templates_dict[new_class], 1, p=template_weights)[0])
        new_type = new_class_stat_dict[new_class]["Type"]

    else:
        new_class = "Failed Adventurer"
        new_template = "Peasant"
        new_type = "Peasant"

    current_char["Weights"] = char_class_weights
    current_char["Class"] = new_class
    current_char["Class Type"] = new_type
    current_char["Experience"] = 0
    current_char["Template"] = new_template
    return current_char


def personal_assign(current_char, index):
    sex = random.choice(sex_list)
    race = current_char.get("Race")

    origin_list = [*new_race_info_dict["Names"][race].keys()]
    origin_weights = [*new_race_info_dict["Names"][race].values()]
    name_origin = str(np.random.choice(origin_list, 1, p=origin_weights)[0])

    first_name = str(random.choice(name_list[name_origin][sex]))
    last_name = str(random.choice(name_list[name_origin]["Last"]))
    risk = int(risk_list[index])
    adventurer = new_race_info_dict["Adventuring"][race][current_char["Level"]] > adv_list[index]

    current_level = current_char.get("Level")
    current_class = current_char.get("Class")

    current_type = current_char.get("Class Type")

    if current_type == "Cleric":
        god_list = [*new_race_info_dict["Gods"][race].keys()]
        god_weights = [*new_race_info_dict["Gods"][race].values()]
        god = str(np.random.choice(god_list, 1, p=god_weights)[0])
    else:
        god = "-"

    # Exp Assign
    if current_class == "Failed Adventurer":
        new_exp = 0
    else:
        level_cap = new_class_stat_dict[current_class]["Experience"]["Cap"]
        if current_level == level_cap:
            new_exp = new_class_stat_dict[current_class]["Experience"][level_cap]
        else:
            min_exp = new_class_stat_dict[current_class]["Experience"][current_level]
            max_exp = new_class_stat_dict[current_class]["Experience"][min(current_level+1, level_cap)] - 1
            new_exp = int(round(np.random.uniform(min_exp, max_exp, 1)[0]))
        current_char["Level"] = min(current_level, level_cap)

    current_char["Experience"] = new_exp
    full_name = first_name + " "+last_name
    if belief_use:
        new_belief = belief_assign(current_class, race, full_name, index)
    else:
        new_belief = ""

    # Update Char Dictionary
    personal_update = {
        "Group": race,
        "Sex": sex,
        "First Name": first_name,
        "Last Name": last_name,
        "Name Origin": name_origin,
        "Risk": risk,
        "God": god,
        "Adventurer": adventurer,
        "Beliefs": new_belief
    }
    return personal_update


def belief_assign(game_class, group, full_name, index):
    report_number = 6
    # full_name = "Stretch Andrews"
    # group = "Ersteland"
    # index = 1
    start_index = index * 33
    end_index = start_index + 33
    current_rolls = belief_rolls[start_index:end_index]
    group_weights = new_race_info_dict["Beliefs"][group]
    class_weights = 0

    belief_values = {}
    for belief, roll in zip(belief_list, current_rolls):
        if belief in group_weights:
            group_mod = group_weights[belief]
        else:
            group_mod = 0
        belief_values[belief] = roll + group_mod

    sorted_belief_dict = dict(sorted(belief_values.items(), key=lambda belief: abs(belief[1]), reverse=True))
    strongest_beliefs = [i for i in sorted_belief_dict if sorted_belief_dict[i] > 10 or sorted_belief_dict[i] < -10][:report_number]

    belief_output = full_name + " "
    for belief in strongest_beliefs:
        current_value = belief_values[belief]
        for cutoff, value in zip(belief_cutoff, belief_index):
            if current_value >= cutoff:
                string_index = value
                break
        belief_string = belief_dict["Beliefs"][belief][string_index]

        if strongest_beliefs.index(belief)+1 == 1:
            belief_output = belief_output + belief_string
        elif strongest_beliefs.index(belief)+1 < len(strongest_beliefs):
            belief_output = belief_output + ", " + belief_string
        else:
            belief_output = belief_output + ", and " + belief_string + "."
    char_belief = {}
    char_belief["Values"] = sorted_belief_dict
    char_belief["Report"] = belief_output
    return char_belief


def final_options():
    option_input = input("Export, Redo, or Exit?\n").lower()
    if option_input == "export":
        export_to_excel()
        final_options()
    elif option_input == "redo":
        return True
    elif option_input == "exit":
        raise SystemExit
    else:
        final_options()


def export_to_excel():
    startTime = time.time()
    print("Exporting data...")
    dest_filename = 'Characters.csv'
    full_path = file_path+'\\'+dest_filename

    char_df.to_csv(full_path, index=False, sep="|")

    # sheet_name = 'Characters'
    # writer = pd.ExcelWriter(full_path, engine='xlsxwriter', mode='w')
    # print("Writer defined.")
    # char_df.style.set_properties(**{'text-align': 'center'}).to_excel(
    #     writer,
    #     index =False,
    #     header =True,
    #     sheet_name = sheet_name
    # )
    # print("Dataframe written.")
    # worksheet = writer.sheets[sheet_name]
    # for column_cells in worksheet.columns:
    #     new_column_length = max(len(str(cell.value)) for cell in column_cells)
    #     new_column_letter = (get_column_letter(column_cells[0].column))
    #     if new_column_length > 0:
    #         worksheet.column_dimensions[new_column_letter].width = new_column_length*1.3

    # writer.close()
    print(f"Export complete!\t[ {(time.time()-startTime):.2f}s ]")


# %%Char Gen Loop
run = True
while run:
    startTime = time.time()
    # %%%Data Fetch
    try:
        if getattr(sys, 'frozen', False):
            file_path = os.path.dirname(sys.executable)
        elif __file__:
            file_path = os.path.dirname(__file__)
    except:
        file_path = os.getcwd()+"\\Desktop\\ACKS-NPC-Generator"

    # ----Excel
    settings_path = file_path+'\\Settings.xlsx'
    settings_file = openpyxl.load_workbook(settings_path, data_only=True)

    stat_prob_dict = excel_dict_fetch(settings_file, 'DictStatProb')
    templates_dict = excel_dict_fetch(settings_file, 'DictTemplates')
    name_list = excel_dict_fetch(settings_file, 'DictNameList')
    new_race_info_dict = excel_dict_fetch(settings_file, 'DictNewRaceInfo')
    new_class_stat_dict = excel_dict_fetch(settings_file, 'DictNewClassInfo')
    belief_dict = excel_dict_fetch(settings_file, 'DictBeliefs')

    # %%%Dictionaries and Lists
    print("Setting dictionaries and lists...")
    # ----Race
    race_list = [race for race in new_race_info_dict["Names"].keys()]
    race_reference = {i: {} for i in race_list}

    # ----Level
    level_result = [-1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

    # ----Stats

    stat_results = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]

    stat_mods = [-3, -2, -2, -1, -1, -1, 0, 0, 0, 0, 1, 1, 1, 2, 2, 3]

    stats = ["STR", "INT", "WIS", "DEX", "CON", "CHA"]
    character_attributes = ["Level", "Race", "Stats", "Class", "Template", "Personal", "Weights"]
    personal_keys = ["Sex", "First Name", "Last Name", "Name Origin", "Risk", "Adventurer", "Beliefs"]

    sex_list = ["M", "F"]

    classes = [i for i in new_class_stat_dict]

    belief_list = [i for i in belief_dict["Beliefs"]]

    # ----Templates

    template_weights = [
        sum(stat_prob_dict['Base'][0:2]),
        sum(stat_prob_dict['Base'][2:4]),
        sum(stat_prob_dict['Base'][4:6]),
        sum(stat_prob_dict['Base'][6:8]),
        sum(stat_prob_dict['Base'][8:10]),
        sum(stat_prob_dict['Base'][10:12]),
        sum(stat_prob_dict['Base'][12:14]),
        sum(stat_prob_dict['Base'][14:16])
    ]
    # ----Beliefs
    belief_cutoff = [41, 26, 11, -10, -25, -40, -50]
    belief_index = [6, 5, 4, 3, 2, 1, 0]

    class character():
        """
        Stuff I won't type right now'
        """

        def __init__(self):
            """"
            """
            self.classname = ""
            self.classtype = ""
            self.exp = 0

        def __repr__(self):
            value = [1, 2]
            return str(value)

        # def parse_level(self,levelroll):
    # %%%Character Generation

    # ---- Creating Initial Dictionary
    current_id = 1
    characters = dict()

    pop_list = new_race_info_dict["Population"]

    print("Group populations provided.")
    print("Beginning character generation...")

    # race_list = list(pop_list.keys())
    # ---- Level Rolling
    for race in race_list:
        pop = pop_list[race]
        race_reference[race]["Total Pop"] = pop
        level_rolls = np.random.choice(level_result, pop, p=[*new_race_info_dict["Level"][race].values()])
        short_level = [i for i in level_rolls if i > 0]
        short_level.sort(reverse=True)
        character_count = len(short_level)

        race_reference[race]["Level"] = short_level
        race_reference[race]["Character Count"] = character_count

        for i in range(character_count):
            current_attributes = {key: None for key in character_attributes}
            current_attributes["Stats"] = {key: None for key in stats}
            current_attributes["Personal"] = {key: None for key in personal_keys}
            current_attributes["Level"] = int(short_level[i])
            characters[current_id] = {i: current_attributes[i] for i in current_attributes}
            characters[current_id]["Race"] = race
            current_id += 1

    # ----Stat Rolls
    total_count = sum([race_reference[race]["Character Count"] for race in race_list])

    stat_types = [new_race_info_dict["Stat Prob"][race][stat] for race in race_list for stat in stats]
    stat_type_count = {key: stat_types.count(key) for key in set(stat_types)}
    stat_rolls = {key: np.random.choice(stat_results, stat_type_count[key] * total_count, p=stat_prob_dict[key]) for key in stat_type_count}

    # ----Personal Rolls
    adv_list = np.random.random(total_count)
    risk_list = np.random.choice(stat_mods, total_count, p=stat_prob_dict['Base'])

    # ----Beliefs

    belief_use = True

    if belief_use:
        belief_range = range(-50, 51, 1)
        belief_weights = belief_dict.get("Belief Weights")
        belief_rolls = list(np.random.choice(belief_range, total_count*33, p=belief_weights))

    # ----Character Assign
    index = 0
    remove_list = []
    stat_index_all = {key: 0 for key in stat_rolls}
    for key in characters:
        current_char = characters.get(key)
        current_char.update(stat_assign(current_char))
        current_char.update(class_assign(current_char))
        current_char["Personal"].update(personal_assign(current_char, index))
        if current_char["Class"] == "Failed Adventurer":
            remove_list.append(key)
        index += 1
        # characters[key].update(current_char)

    for key in remove_list:
        characters.pop(key)

    for key in characters:
        if characters[key]["Class"] == "Failed Adventurer":
            print(key)

    # ----Data Prep

    headers = [
        "ID", "STR", "INT", "WIS", "DEX", "CON", "CHA", "V",
        "Group", "Name", "Class", "Type", "Sex", "Level",
        "Exp", "Stats", "Risk", "Template", "God", "Adventurer", "Personality", "Beliefs"
    ]

    spreadsheet_export = [headers]
    belief_export = {i: [] for i in ["Class"] + belief_list}
    excel_export = {i: [] for i in headers}

    for key in characters:
        char = characters.get(key)
        stats = char.get("Stats")
        per = char.get("Personal")
        # adv_check = per.get("Adventurer")
        adv_check = True

        if adv_check:
            STR = stats.get("STR")
            INT = stats.get("INT")
            WIS = stats.get("WIS")
            DEX = stats.get("DEX")
            CON = stats.get("CON")
            CHA = stats.get("CHA")
            stat_value = sum(stats.values())
            name = per.get("First Name")+" "+per.get("Last Name")
            game_class = char.get("Class")
            sex = per.get("Sex")
            level = char.get("Level")
            exp = char.get("Experience")
            stat_list = "/".join([str(i) for i in stats.values()])
            risk = per.get("Risk")
            template = char.get("Template")
            god = per.get("God")
            adv = per.get("Adventurer")
            if adv:
                adv = "TRUE"
            else:
                adv = "FALSE"
            group = per.get("Group")
            class_type = char.get("Class Type")

            if belief_use:
                top_beliefs = per.get("Beliefs").get("Report")
                belief_values = per.get("Beliefs").get("Values")
                belief_export["Class"].append(game_class)
                for belief in belief_list:
                    belief_export[belief].append(belief_values.get(belief))
            else:
                top_beliefs = ""
                belief_values = ""

            new_entry = [key, STR, INT, WIS, DEX, CON, CHA, stat_value, group, name, game_class, class_type, sex, level, exp, stat_list, risk, template, god, adv, top_beliefs, belief_values]
            spreadsheet_export.append(new_entry)
            for i in range(len(headers)):
                excel_export[headers[i]].append(new_entry[i])
    print(f"Character generation complete!\t[ {(time.time()-startTime):.2f}s ]")
    # %%% Analysis
    analysis_check = True
    if analysis_check:
        startTime = time.time()
        # ----Dataframe Creation
        char_df = pd.DataFrame(excel_export).sort_values(["Level", "Exp", "Name"], ascending=[False, False, True])
        # ----Class Numbers
        class_list = [characters[i]["Class"] for i in characters]
        unique_class = set(class_list)
        class_count = [class_list.count(i) for i in list(new_class_stat_dict.keys())]
        class_numbers = {classes[i]: class_count[i] for i in range(len(list(new_class_stat_dict.keys())))}
        # ----Class Search
        search = []
        for i in characters:
            char_class = characters[i]["Class"]
            if char_class == "Sage":
                search.append(i)

        # ----Class Stat Averages
        stat_analysis = char_df.groupby(["Group", "Class"]).agg(
            {
                "STR": np.mean,
                "INT": np.mean,
                "WIS": np.mean,
                "DEX": np.mean,
                "CON": np.mean,
                "CHA": np.mean,
            }

        ).round(2)
        # ----Group/Class Breakdown
        group_1 = pd.DataFrame({"Percent": (1 / char_df.groupby(["Group"]).size())*100})
        group_type = pd.merge(char_df, group_1, on=["Group"])
        group_analysis = group_type.groupby(["Group", "Type", "Class"])
        group_analysis = group_analysis.agg(
            Count=pd.NamedAgg(column="Class", aggfunc="count"),
            Percent=pd.NamedAgg(column="Percent", aggfunc="sum")
        )
        group_analysis = group_analysis.round(2).sort_values(["Group", "Type", "Count"], ascending=[False, True, False])

        # ----Class Type Breakdown
        class_type = char_df.assign(Percent=(1 / char_df["ID"].count())*100).groupby("Type").agg({"Percent":  np.sum}).round(2).sort_values(["Percent"], ascending=False)

        # ----Belief Spread
        # belief_df = pd.DataFrame(belief_export)
        # belief_analysis = belief_df.groupby(["Class"]).agg('mean').astype(int)

        print(stat_analysis)
        print(group_analysis)
        print(class_type)
        # print(belief_analysis)
        print(f"Analysis complete.\t[ {(time.time()-startTime):.2f}s ]")
    # %%%Final Options
    final_options()


# for i in range(-50,51,1):
#     print(i)
# %%Dead Code

# workbook = xlsxwriter.Workbook('pandas_simple.xlsx')
# cell_format = workbook.add_format().set_align('center')
# #worksheet.set_column('A:S',None, cell_format)
# new_rows = len(char_df)+1
# new_cols = len(new_entry)
# start_range = "A1"
# end_range = xlref(new_rows,new_cols,False)
# dest_filename = 'empty_book.xlsx'
# full_path = file_path+'/'+dest_filename
# wb = load_workbook(filename = full_path)
# ws = wb.active
# cell_range = ws[start_range : end_range]
# for r in dataframe_to_rows(char_df, index=False, header=True):
#     ws.append(r)
# for cell in ws[1]:
#     cell.style = 'Pandas'
# wb.save(full_path)
# char_df.to_excel('./test.xlsx', sheet_name='Characters', index=False)

# level_rolls = np.random.random(population)

# level_search = [0,0.86805555,0.95809750,0.98476132,0.99445726,0.99798305,0.99926516,0.99973138,0.99990091,0.99996256,0.99998498,0.99999395,0.99999754,0.99999913,0.99999971,0.99999992]
# level_result = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14]

# level_list = [check_lowest(level_search,level_result,i) for i in level_rolls]


# analysis_percent = char_df.groupby(["Group","Class"]).agg(Count=pd.NamedAgg(column="Class", aggfunc="count"))
# analysis_percent = analysis_percent.assign(Percent=(analysis_percent["Count"] / char_df.groupby("Group")["ID"].count())*100)
# analysis_percent = analysis_percent.round(2).sort_values(["Group","Percent"], ascending=False)
